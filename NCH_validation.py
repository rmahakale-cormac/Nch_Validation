#open execute this program for all file in a folder
import os
from datetime import date
from datetime import datetime
import smtplib
import shutil
import pandas as pd
from openpyxl import load_workbook
#email variables
recev1 = "rajatmahakale@gmail.com"
myemail = "nacharchival@gmail.com"
passwd = "Ncharchival123"
body =""

#file parameters
f_path = "/home/cd_admin/NCH_Archival_Validation/"
s_path = "/home/cd_admin/NCH_Archival_Validation/Successful/"
u_path = "/home/cd_admin/NCH_Archival_Validation/Unsuccessful/"
f_ext =".*"
#-----------------------------------------------------------------------

today = date.today()
time = datetime.now()
value10,amt10, price,value81, value71,amt71,value72,amt72,value82,amt81,amt82,value50,amt50,=0,0,0,0,0,0,0,0,0,0,0,0,0

#search all the files ending with .txt
#file_names = [x for x in os.listdir(f_path) if x.endswith(f_ext)]
file_names = ['D#UTL.#NCHP.NLK.A6.Y15.S03BX.CONV']
#print(file_names);

#read all the files in the folder
for names in file_names:
   # os.chdir('/home/cd_admin/NCH_Archival_Validation/files/')
    fileHandle = open(names,'r')
    lineList = fileHandle.readlines()
    last_line=(lineList[-1])

    print(last_line)
    #generate data dictionary here
    file = open(names, 'r')
    lines = file.readlines()
    filesize = os.path.getsize(names)
    filesize= (filesize)/1000

    for line in lines[0:len(lines)-1]:
        a = line
        awsmetadata=""
        claim_type = a[10:11]
        if claim_type == "72" or "71" or "81" or "82":
            paid_amt=a[242:254]
        else:
            paid_amt = a[245:257]
        # print(paid_amt)
        intaamt = int(paid_amt)

        if claim_type == "010":
            value10 += 1
            aamt = (aamt + intaamt)

        if claim_type == "71":
            value71 += 1
            intamt71 = int(paid_amt)
            amt71 = (amt71 + intamt71)
        if claim_type == "72":
            value72 += 1
            intamt72 = int(paid_amt)
            amt72 = (amt72 + intamt72)
        if claim_type == "81":
            value81 += 1
            intamt81 = int(paid_amt)
            amt72 = (amt81 + intamt81)
        if claim_type == "82":
            value82 += 1
            intamt82 = int(paid_amt)
            amt72 = (amt82 + intamt82)
        if claim_type == "50":
            value50 += 1
            intamt50 = int(paid_amt)
            amt72 = (amt50 + intamt50)

awsmetadata = (str(value10) + ";" + str(amt10) + ";" + str(value71) + ";" + str(amt71) + ";" + str(value72) + ";" + str(amt72)+ str(value81) + ";" + str(amt81)+ str(value82) + ";" + str(amt82)+ str(value50) + ";" + str(amt50))
value10,amt10, price,value81, value71,amt71,value72,amt72,value82,amt81,amt82,value50,amt50,=0,0,0,0,0,0,0,0,0,0,0,0,0
# print(awsmetadata)


# Compare metadata here
comparision = (awsmetadata.replace(" ", "") == (last_line.replace(" ", "")))

if comparision == True:
    print("Validation SUCCEEDED for file " + names + " of size " + str(filesize) + " KB at", time)
    df = pd.DataFrame({'File Name': [names],
                       'Validation': ['Successful'],
                       'Size': [filesize],
                       'Time': [time]})
    # os.chdir('/home/cd_admin/NCH_Archival_Validation/')
    writer = pd.ExcelWriter('Report.xlsx', engine='openpyxl')
    # writer.book = load_workbook('Report.xlsx')
    # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # reader = pd.read_excel(r'Report.xlsx')
    df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)

    writer.save()
    writer.close()
    file.close()
    fileHandle.close()
    filewrite = open(names, 'w')
    del lineList[-1]
    for newlines in lineList:
        filewrite.write(newlines)
    filewrite.close()
    newname = names.replace('.txt', '')
    newname = (newname + 'AWS.txt')
    os.rename(names, newname)
    # os.chdir('/home/cd_admin/NCH_Archival_Validation/files/')
    # shutil.move(newname,s_path)

elif comparision == False:
    print("Validation FAILED for file " + names + " of size " + str(filesize) + " KB at", time)
    df = pd.DataFrame({'File Name': [names],
                       'Validation': ['Failed'],
                       'Size': [filesize],
                       'Time': [time]})

    # Generate reports here
    appendbody = ("\nValidation FAILED for file " + names + " of size " + str(filesize) + " bytes")
    body = str(body) + str(appendbody)
    # shutil.move(names, u_path )   #delete file here
    # os.chdir('/home/cd_admin/NCH_Archival_Validation/')
    writer = pd.ExcelWriter('Report.xlsx', engine='openpyxl')
    writer.book = load_workbook('Report.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(r'Report.xlsx')
    df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
    writer.save()
    writer.close()
    file.close()
    fileHandle.close()

# _Send email here
to = [recev1]
gmail_user = myemail
gmail_password = passwd
subject = 'NCH Archival File Validation Failed'

email_text = """\
    From: %s
    To: %s
    Subject: %s

    %s
    """ % (gmail_user, ", ".join(to), subject, body)
server_ssl = smtplib.SMTP_SSL('smtp.gmail.com', 465)
server_ssl.ehlo()
# server_ssl.login(gmail_user, gmail_password)
# if body:
# server_ssl.sendmail(gmail_user, to, email_text)#--------------Remove once code is migrated
